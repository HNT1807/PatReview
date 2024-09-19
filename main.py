import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from streamlit.components.v1 import html
import json



# Initialize session state variables
if 'reset_triggered' not in st.session_state:
    st.session_state.reset_triggered = False
if 'file_uploader_key' not in st.session_state:
    st.session_state.file_uploader_key = 0

def trigger_reset():
    st.session_state.reset_triggered = True

def reset_app():
    for key in list(st.session_state.keys()):
        if key != 'reset_triggered' and key != 'file_uploader_key':
            del st.session_state[key]
    st.session_state.modified_cells = set()
    st.session_state.df = None
    st.session_state.history = []
    st.session_state.history_index = -1
    st.session_state.new_text = ""
    st.session_state.track_title = ""
    st.session_state.column_letters = ""
    st.session_state.find_text = ""
    st.session_state.file_uploader_key += 1  # Increment the key to force a new file uploader
    st.session_state.reset_triggered = False



if st.session_state.reset_triggered:
    reset_app()
st.set_page_config(layout="wide")

def save_state():
    st.session_state.history = st.session_state.history[:st.session_state.history_index + 1]
    st.session_state.history.append(st.session_state.df.copy())
    st.session_state.history_index += 1


# Place this at the beginning of your script, after imports and before any other code
if st.session_state.get('reset_requested', False):
    st.session_state.reset_requested = False
    reset_app()





st.markdown("""
    <style>
    button[data-testid="download-button"][aria-label="DOWNLOAD XL"] {
        background-color: #4CAF50;
        color: white;
    }
    </style>
""", unsafe_allow_html=True)


def download_excel():
    if st.session_state.df is not None and not st.session_state.df.empty:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            st.session_state.df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Apply highlighting to modified cells
            for col in st.session_state.modified_cells:
                for row in range(len(st.session_state.df)):
                    cell = worksheet.cell(row=row + 2, column=col + 1)  # row + 2 because Excel is 1-indexed and we have a header
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

        output.seek(0)
        return output.getvalue()
    return None


def highlight_modified_cells(df):
    def highlight(val):
        return 'background-color: #ADD8E6'

    styled = df.style.applymap(lambda _: '')  # Initialize all cells with no style
    if 'modified_cells' in st.session_state and st.session_state.modified_cells:
        for col in st.session_state.modified_cells:
            if col < len(df.columns):
                styled = styled.applymap(highlight, subset=pd.IndexSlice[:, df.columns[col]])
    return styled

# Add this near the top of your script, where other session state variables are initialized
if 'first_updated_cell' not in st.session_state:
    st.session_state.first_updated_cell = None

# Modify the update_cell function
def update_cell(row, col, value):
    if st.session_state.df.iloc[row, col] != value:
        st.session_state.df.iloc[row, col] = value
        if 'modified_cells' not in st.session_state:
            st.session_state.modified_cells = set()
        st.session_state.modified_cells.add(col)
        if st.session_state.first_updated_cell is None:
            st.session_state.first_updated_cell = (row, col)
        print(f"Cell updated: row {row}, col {col}, value {value}")
        print(f"Modified cells: {st.session_state.modified_cells}")
# Add this JavaScript function to scroll to a specific cell
st.markdown("""
<script>
function scrollToCell(row, col) {
    const dataFrame = document.querySelector('.stDataFrame');
    if (dataFrame) {
        const table = dataFrame.querySelector('table');
        if (table) {
            const cell = table.rows[row + 1].cells[col];  // +1 because of header row
            if (cell) {
                cell.scrollIntoView({ behavior: 'smooth', block: 'center' });
                cell.focus();
            }
        }
    }
}
</script>
""", unsafe_allow_html=True)


def apply_subcategories(df, new_sub_category, track_title=None):
    subcategory_column = 'SubCategory'
    track_title_column = 'TrackTitle'

    if subcategory_column not in df.columns:
        st.error(f"Column '{subcategory_column}' not found in the DataFrame.")
        return df

    if track_title:
        # Apply the new subcategory to non-empty cells in the subcategory column where track title matches (case-insensitive)
        mask = (df[subcategory_column].notna()) & (df[track_title_column].str.lower() == track_title.lower())
    else:
        # Apply the new subcategory to all non-empty cells in the subcategory column
        mask = df[subcategory_column].notna()

    df.loc[mask, subcategory_column] = new_sub_category

    # Add the modified column to the set of modified cells
    if 'modified_cells' not in st.session_state:
        st.session_state.modified_cells = set()
    st.session_state.modified_cells.add(df.columns.get_loc(subcategory_column))

    return df


def apply_cd_description(df, album_description):
    # Assuming column 'Q' is the 17th column (index 16)
    column_index = 16

    # Apply the album description to non-empty cells in column 'Q'
    mask = df.iloc[:, column_index].notna()
    df.loc[mask, df.columns[column_index]] = album_description

    return df


def apply_to_all_cells_in_column(df, column_letters, new_text):
    if not column_letters or not new_text:
        return df

    cols = [ord(c.upper()) - 65 for c in column_letters.split(',') if c.strip()]

    for col in cols:
        if col < 0 or col >= len(df.columns):
            st.warning(f"Column {chr(col + 65)} is out of range and will be skipped.")
            continue

        df.iloc[:, col] = new_text

        # Add the modified column to the set of modified cells
        if 'modified_cells' not in st.session_state:
            st.session_state.modified_cells = set()
        st.session_state.modified_cells.add(col)

    return df


def add_text_before_in_column(df, column_letters, new_text, track_title=None):
    if not column_letters or not new_text:
        return df

    cols = [ord(c.upper()) - 65 for c in column_letters.split(',') if c.strip()]

    for col in cols:
        if col < 0 or col >= len(df.columns):
            continue

        if track_title:
            # Case-insensitive match for track title, using 'TrackTitle' column
            mask = df['TrackTitle'].astype(str).str.lower() == track_title.lower()
            if mask.any():
                df.loc[mask, df.columns[col]] = new_text + ", " + df.loc[mask, df.columns[col]].astype(str)
            else:
                st.warning(f"No rows found matching track title: {track_title}")
        else:
            df.iloc[:, col] = new_text + ", " + df.iloc[:, col].astype(str)

        # Add the modified column to the set of modified cells
        if 'modified_cells' not in st.session_state:
            st.session_state.modified_cells = set()
        st.session_state.modified_cells.add(col)

    return df

def add_text_after_in_column(df, column_letters, new_text, track_title=None):
    if not column_letters or not new_text:
        return df

    cols = [ord(c.upper()) - 65 for c in column_letters.split(',') if c.strip()]

    for col in cols:
        if col < 0 or col >= len(df.columns):
            continue

        if track_title:
            # Case-insensitive match for track title, using 'TrackTitle' column
            mask = df['TrackTitle'].astype(str).str.lower() == track_title.lower()
            if mask.any():
                df.loc[mask, df.columns[col]] = df.loc[mask, df.columns[col]].astype(str) + ", " + new_text
            else:
                st.warning(f"No rows found matching track title: {track_title}")
        else:
            df.iloc[:, col] = df.iloc[:, col].astype(str) + ", " + new_text

        # Add the modified column to the set of modified cells
        if 'modified_cells' not in st.session_state:
            st.session_state.modified_cells = set()
        st.session_state.modified_cells.add(col)

    return df




def find_and_replace_in_column(df, column_letters, find_text, new_text, track_title=None):
    if not column_letters or not find_text or not new_text:
        return df

    cols = [ord(c.upper()) - 65 for c in column_letters.split(',') if c.strip()]

    for col in cols:
        if col < 0 or col >= len(df.columns):
            st.warning(f"Column {chr(col + 65)} is out of range and will be skipped.")
            continue

        if track_title:
            # Case-insensitive match for track title, using 'TrackTitle' column
            mask = df['TrackTitle'].astype(str).str.lower() == track_title.lower()
            if mask.any():
                df.loc[mask, df.columns[col]] = df.loc[mask, df.columns[col]].astype(str).str.replace(find_text, new_text, case=False, regex=False)
            else:
                st.warning(f"No rows found matching track title: {track_title}")
        else:
            df.iloc[:, col] = df.iloc[:, col].astype(str).str.replace(find_text, new_text, case=False, regex=False)

        # Add the modified column to the set of modified cells
        if 'modified_cells' not in st.session_state:
            st.session_state.modified_cells = set()
        st.session_state.modified_cells.add(col)

    return df


# Initialize session state
if 'modified_cells' not in st.session_state:
    st.session_state.modified_cells = set()
if 'df' not in st.session_state:
    st.session_state.df = None
if 'history' not in st.session_state:
    st.session_state.history = []
if 'history_index' not in st.session_state:
    st.session_state.history_index = -1

# Check for reset request
if st.session_state.get('reset_requested', False):
    st.session_state.reset_requested = False
    reset_app()

st.markdown("""
    <style>
    div.stDownloadButton > button:first-child {
        background-color: #4CAF50;
        color: white;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
    <style>
    div.stButton > button {
        width: 100%;
    }
    div.st-emotion-cache-1kyxreq {
        display: flex;
        justify-content: center;
        align-items: center;
    }
    </style>
""", unsafe_allow_html=True)

# Sidebar for controls
with st.sidebar:
    st.markdown("<h1 style='text-align: center;'>CONTROLS</h1>", unsafe_allow_html=True)

    uploaded_file = st.sidebar.file_uploader("", type=["xlsx", "xls"],
                                             key=f"uploaded_file_{st.session_state.file_uploader_key}")

    # After the file uploader
    if uploaded_file is not None:
        if 'df' not in st.session_state or st.session_state.df is None:
            st.session_state.df = pd.read_excel(uploaded_file)
            save_state()

        file_name = uploaded_file.name
        pw_file_name = f"{file_name.rsplit('.', 1)[0]}.PW.xlsx"
        st.text_input("File name for the updated XL:", value=pw_file_name, key="file_name")

        excel_data = download_excel()
        if excel_data is not None:
            if st.download_button(
                label="DOWNLOAD XL",
                data=excel_data,
                file_name=st.session_state.file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Click to download the modified Excel file",
                use_container_width=True,
            ):
                st.balloons()
        else:
            st.write("No data available for download. Please check if the file was uploaded correctly.")

        col1, col2, col3 = st.columns([1, 0.2, 1])
        with col1:
            if st.button("UNDO", use_container_width=True):
                if st.session_state.history_index > 0:
                    st.session_state.history_index -= 1
                    st.session_state.df = st.session_state.history[st.session_state.history_index].copy()
        with col3:
            if st.button("REDO", use_container_width=True):
                if st.session_state.history_index < len(st.session_state.history) - 1:
                    st.session_state.history_index += 1
                    st.session_state.df = st.session_state.history[st.session_state.history_index].copy()

    new_text = st.text_input("NEW TEXT:", key="new_text", value=st.session_state.get('new_text', ''))
    track_title = st.text_input("TRACK TITLE:", key="track_title", value=st.session_state.get('track_title', ''))
    column_letters = st.text_input("COLUMN LETTER(S):", key="column_letters",
                                   value=st.session_state.get('column_letters', ''))
    find_text = st.text_input("FIND TEXT:", key="find_text", value=st.session_state.get('find_text', ''))

    # Button layout
    st.write("")  # Add some space
    container = st.container()

    with container:
        col1, col2 = st.columns(2)
        with col1:
            if st.button("SUBCATEGORIES", use_container_width=True):
                if st.session_state.df is not None:
                    new_sub_category = st.session_state.get('new_text', '')
                    track_title = st.session_state.get('track_title', '')
                    st.session_state.df = apply_subcategories(st.session_state.df, new_sub_category,
                                                              track_title if track_title else None)
                    save_state()
                else:
                    st.error("Please upload an Excel sheet first.")

        with col2:
            if st.button("CD DESCRIPTION", use_container_width=True):
                if st.session_state.df is not None:
                    album_description = st.session_state.get('new_text', '')
                    st.session_state.df = apply_cd_description(st.session_state.df, album_description)
                    save_state()
                else:
                    st.error("Please upload an Excel sheet first.")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("ADD TEXT IN COLUMN (BEFORE)", use_container_width=True):
                if st.session_state.df is not None:
                    column_letters = st.session_state.get('column_letters', '')
                    new_text = st.session_state.get('new_text', '')
                    track_title = st.session_state.get('track_title', '')
                    if column_letters and new_text:
                        original_df = st.session_state.df.copy()
                        st.session_state.df = add_text_before_in_column(st.session_state.df, column_letters, new_text,
                                                                        track_title)
                        if not st.session_state.df.equals(original_df):
                            save_state()
                        else:
                            st.warning("No changes were made to the DataFrame.")
                    else:
                        st.error("Please specify both column letter(s) and new text.")
                else:
                    st.error("Please upload an Excel sheet first.")

        with col2:
            if st.button("ADD TEXT IN COLUMN (AFTER)", use_container_width=True):
                if st.session_state.df is not None:
                    column_letters = st.session_state.get('column_letters', '')
                    new_text = st.session_state.get('new_text', '')
                    track_title = st.session_state.get('track_title', '')
                    if column_letters and new_text:
                        original_df = st.session_state.df.copy()
                        st.session_state.df = add_text_after_in_column(st.session_state.df, column_letters, new_text,
                                                                       track_title)
                        if not st.session_state.df.equals(original_df):
                            save_state()
                        if track_title:
                            st.success(
                                f"Added ', {new_text}' after existing text in specified column(s): {column_letters} for track title: {track_title}")
                        else:
                            st.success(
                                f"Added ', {new_text}' after existing text in specified column(s): {column_letters} for all tracks")
                    else:
                        st.error("Please specify both column letter(s) and new text.")
                else:
                    st.error("Please upload an Excel sheet first.")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("FIND AND REPLACE IN COLUMN", use_container_width=True):
                if st.session_state.df is not None:
                    column_letters = st.session_state.get('column_letters', '')
                    find_text = st.session_state.get('find_text', '')
                    new_text = st.session_state.get('new_text', '')
                    track_title = st.session_state.get('track_title', '')
                    if column_letters and find_text and new_text:
                        original_df = st.session_state.df.copy()
                        st.session_state.df = find_and_replace_in_column(st.session_state.df, column_letters, find_text,
                                                                         new_text, track_title)
                        if not st.session_state.df.equals(original_df):
                            save_state()
                        else:
                            st.warning("No changes were made to the DataFrame.")
                    else:
                        st.error("Please specify column letter(s), find text, and new text.")
                else:
                    st.error("Please upload an Excel sheet first.")

        with col2:
            if st.button("ALL CELLS IN COLUMN", use_container_width=True):
                if st.session_state.df is not None:
                    column_letters = st.session_state.get('column_letters', '')
                    new_text = st.session_state.get('new_text', '')
                    if column_letters and new_text:
                        st.session_state.df = apply_to_all_cells_in_column(st.session_state.df, column_letters,
                                                                           new_text)
                        save_state()
                    else:
                        st.error("Please specify both column letter(s) and new text.")
                else:
                    st.error("Please upload an Excel sheet first.")

    # Reset button
    st.button("RESET APP", on_click=trigger_reset, key="reset-app-button", help="Click to reset the app",
              use_container_width=True)






# Main area for preview
if uploaded_file is not None:
    if st.session_state.df is not None:
        # Use the full height of the screen for the dataframe
        st.markdown("""
            <style>
            .main .block-container {
                padding-top: 1rem;
                padding-bottom: 1rem;
                height: 100vh;
            }
            .stDataFrame {
                height: calc(100vh - 2rem);
            }
            .stDataFrame > div {
                height: 100% !important;
            }
            </style>
            <script>
            function adjustDataFrameHeight() {
                const dataFrame = document.querySelector('.stDataFrame > div');
                if (dataFrame) {
                    const windowHeight = window.innerHeight;
                    const dataFrameTop = dataFrame.getBoundingClientRect().top;
                    const newHeight = windowHeight - dataFrameTop - 20; // 20px for some bottom margin
                    dataFrame.style.height = `${newHeight}px`;
                }
            }
            // Run on load and whenever the window is resized
            window.addEventListener('load', adjustDataFrameHeight);
            window.addEventListener('resize', adjustDataFrameHeight);
            // Run periodically to catch any layout changes
            setInterval(adjustDataFrameHeight, 1000);
            </script>
        """, unsafe_allow_html=True)

        # Create a styled dataframe
        styled_df = highlight_modified_cells(st.session_state.df)

        # Display the styled dataframe with data_editor
        edited_df = st.data_editor(
            styled_df,
            num_rows="dynamic",
            key="data_editor",
            disabled=("column_config"),
            hide_index=True,
            use_container_width=True,
            height=600  # Set an initial height, it will be adjusted by JavaScript
        )

        # Check for changes and update the session state
        if not st.session_state.df.equals(edited_df):
            for i in range(len(st.session_state.df)):
                for j in range(len(st.session_state.df.columns)):
                    if st.session_state.df.iloc[i, j] != edited_df.iloc[i, j]:
                        update_cell(i, j, edited_df.iloc[i, j])
            st.session_state.df = edited_df.copy()
            save_state()

            # Scroll to the first updated cell
            if st.session_state.first_updated_cell:
                row, col = st.session_state.first_updated_cell
                st.components.v1.html(
                    f'<script>scrollToCell({row}, {col});</script>',
                    height=0,
                    width=0,
                )
                st.session_state.first_updated_cell = None  # Reset for next update
    else:
        st.write("Please upload an Excel sheet to view it.")
else:
    st.write("Please upload an Excel sheet to view it.")